"""
Dataset D Formatter — v1.7.1
True Bearing LLC → IC Sentinel → CIMScan

Post-processes any generated Dataset D .xlsx to apply professional formatting.
Run after outputBuilder writes the file, before upload to storage.

Usage:
    python format_dataset_d.py <input.xlsx> [output.xlsx]

If output is omitted, formats in place.

Exit codes:
    0  Success
    1  File not found or unreadable
    2  Not a valid Dataset D workbook (missing expected sheets)
    3  Formatting failure
"""

import sys
import os
import copy
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

# ─── Palette (matches IC Insights template) ───────────────────────

NAVY       = "1B2A4A"
GOLD       = "C4933F"
DARK_GRAY  = "333333"
MED_GRAY   = "666666"
LIGHT_BG   = "F5F6F8"
TABLE_HDR  = "1F3864"  # existing header fill — keep for continuity
WHITE      = "FFFFFF"
RULE_LINE  = "D0D3DA"

# Conditional formatting colors
RED_BG     = "FDE8E8"
RED_TEXT    = "991B1B"
AMBER_BG   = "FEF3CD"
AMBER_TEXT  = "856404"
GREEN_BG   = "D4EDDA"
GREEN_TEXT  = "155724"
BLUE_BG    = "DBEAFE"
BLUE_TEXT   = "1E40AF"
PURPLE_BG  = "EDE9FE"
PURPLE_TEXT = "5B21B6"

# Hub tag color map
HUB_COLORS = {
    "HUB: Retention Kill":       (RED_BG, RED_TEXT),
    "HUB: Margin Kill":          (AMBER_BG, AMBER_TEXT),
    "HUB: Revenue Concentration": (RED_BG, RED_TEXT),
    "HUB: Multiple Support":     (BLUE_BG, BLUE_TEXT),
    "HUB: Compliance Close Risk": (PURPLE_BG, PURPLE_TEXT),
}

# ─── Style factories ─────────────────────────────────────────────

HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=TABLE_HDR)
HEADER_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Arial", size=10, color=DARK_GRAY)
DATA_FONT_BOLD = Font(name="Arial", size=10, color=DARK_GRAY, bold=True)
DATA_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
DATA_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="top")

ALT_ROW_FILL = PatternFill("solid", fgColor=LIGHT_BG)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)

THIN_BORDER_SIDE = Side(style="thin", color=RULE_LINE)
CELL_BORDER = Border(
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
)

# ─── README styling ──────────────────────────────────────────────

README_TITLE_FONT = Font(name="Arial", size=18, bold=True, color=NAVY)
README_SUBTITLE_FONT = Font(name="Arial", size=11, color=MED_GRAY, italic=True)
README_LABEL_FONT = Font(name="Arial", size=10, bold=True, color=NAVY)
README_VALUE_FONT = Font(name="Arial", size=10, color=DARK_GRAY)
README_SECTION_FONT = Font(name="Arial", size=10, bold=True, color=GOLD)
GOLD_FILL = PatternFill("solid", fgColor=GOLD)


def style_readme(ws):
    """Restyle the README sheet with IC Insights branding."""
    # Clear existing formatting
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.font = README_VALUE_FONT
            cell.fill = WHITE_FILL
            cell.border = Border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 65

    # Row 1: title row — merge and style
    ws.merge_cells("A1:B1")
    ws["A1"].value = "Dataset D"
    ws["A1"].font = README_TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36

    # Row 2: subtitle
    ws.merge_cells("A2:B2")
    ws["A2"].value = "EC-CIM Analytical Workbook"
    ws["A2"].font = README_SUBTITLE_FONT
    ws.row_dimensions[2].height = 20

    # Row 3: gold accent line (empty row with bottom border)
    ws.merge_cells("A3:B3")
    ws["A3"].value = None
    ws["A3"].border = Border(bottom=Side(style="medium", color=GOLD))
    ws["B3"].border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[3].height = 8

    # Rows 4+: metadata and sheet listing
    # Re-populate cleanly
    meta_rows = []

    # Find existing data
    existing = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
        if row[0] and row[1]:
            existing[str(row[0]).strip()] = str(row[1]).strip()

    version = existing.get("EC-CIM Version", "1.7.0")
    entity = existing.get("Entity", "True Bearing LLC > IC Sentinel > CIMScan")
    date = existing.get("Date", "")
    changes = existing.get("v1.7.0 Changes", existing.get("v1.7.1 Changes", ""))
    invariants = existing.get("Governing Invariants", "")
    errors = existing.get("Error Codes", "")

    info_block = [
        ("EC-CIM Version", version),
        ("Entity", entity),
        ("Date", date),
    ]

    sheet_block = []
    for i, name in enumerate(ws.parent.sheetnames):
        if name == "README":
            continue
        sheet_block.append((f"Sheet {i}", name))

    notes_block = []
    if changes:
        notes_block.append(("Changes", changes))
    if invariants:
        notes_block.append(("Governing Invariants", invariants))
    if errors:
        notes_block.append(("Error Codes", errors))

    # Write info block starting at row 5
    row_num = 5
    for label, value in info_block:
        ws.cell(row=row_num, column=1, value=label).font = README_LABEL_FONT
        ws.cell(row=row_num, column=2, value=value).font = README_VALUE_FONT
        ws.cell(row=row_num, column=1).alignment = Alignment(vertical="top")
        ws.cell(row=row_num, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        row_num += 1

    # Spacer
    row_num += 1

    # Sheet index header
    ws.cell(row=row_num, column=1, value="SHEET INDEX").font = README_SECTION_FONT
    ws.cell(row=row_num, column=2, value="").font = README_SECTION_FONT
    ws.cell(row=row_num, column=1).border = Border(bottom=Side(style="thin", color=GOLD))
    ws.cell(row=row_num, column=2).border = Border(bottom=Side(style="thin", color=GOLD))
    row_num += 1

    for label, value in sheet_block:
        ws.cell(row=row_num, column=1, value=label).font = README_LABEL_FONT
        ws.cell(row=row_num, column=2, value=value).font = README_VALUE_FONT
        # alternating bg
        if row_num % 2 == 0:
            ws.cell(row=row_num, column=1).fill = ALT_ROW_FILL
            ws.cell(row=row_num, column=2).fill = ALT_ROW_FILL
        row_num += 1

    # Spacer
    row_num += 1

    # Notes
    if notes_block:
        ws.cell(row=row_num, column=1, value="NOTES").font = README_SECTION_FONT
        ws.cell(row=row_num, column=1).border = Border(bottom=Side(style="thin", color=GOLD))
        ws.cell(row=row_num, column=2).border = Border(bottom=Side(style="thin", color=GOLD))
        row_num += 1

        for label, value in notes_block:
            ws.cell(row=row_num, column=1, value=label).font = README_LABEL_FONT
            ws.cell(row=row_num, column=2, value=value).font = README_VALUE_FONT
            ws.cell(row=row_num, column=2).alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row_num].height = None  # auto
            row_num += 1

    # Clean any leftover rows
    for r in range(row_num, ws.max_row + 1):
        for c in range(1, 3):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.font = README_VALUE_FONT
            cell.fill = WHITE_FILL


# ─── Data sheet styling ──────────────────────────────────────────

def find_header_map(ws):
    """Return {header_name: col_index} from row 1."""
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[str(val).strip()] = c
    return headers


def col_letter(col_idx):
    return get_column_letter(col_idx)


# Suggested widths per column name (overrides for readability)
COLUMN_WIDTHS = {
    "claim_id": 14,
    "claim_id_a": 14,
    "claim_id_b": 14,
    "hub_claim_id": 14,
    "claim_text": 55,
    "claim_category": 20,
    "cim_section": 18,
    "claim_priority_score": 12,
    "source_page": 10,
    "source_excerpt": 45,
    "mechanism_of_value": 40,
    "economic_driver": 14,
    "kpi_to_validate": 25,
    "claim_type": 14,
    "underwriting_gate": 45,
    "kill_threshold": 40,
    "downside_case_if_false": 45,
    "gate_status": 12,
    "diligence_task_type": 22,
    "artifact_name": 30,
    "interview_target": 25,
    "purpose": 45,
    "kpi_shared": 10,
    "driver_shared": 10,
    "semantic": 10,
    "evidence_chain": 10,
    "relationship_strength": 12,
    "relationship_type": 22,
    "blast_radius": 12,
    "linked_claims": 40,
    "hub_classification_tag": 26,
    "hub_tag": 26,
    "cascade_claims": 40,
    "propagation_chain": 45,
    "coupling_family": 18,
    "dimension_signature": 20,
    "inconsistency_description": 45,
    "severity": 12,
    "rank": 8,
    "ic_gating_rationale": 50,
    "pillar_id": 10,
    "pillar_name": 30,
    "pillar_thesis": 50,
    "supporting_claim_ids": 30,
    "key_kpis": 25,
    "hub_claims_linked": 25,
    "blast_radius_exposure": 12,
    "negative_coupling_trigger": 40,
    "pillar_collapse_path_if_breached": 45,
    "ic_red_threshold_condition": 40,
    "Check": 30,
    "Result": 14,
    "Detail": 50,
}

# Columns that should be center-aligned
CENTER_COLS = {
    "claim_priority_score", "source_page", "economic_driver", "claim_type",
    "gate_status", "kpi_shared", "driver_shared", "semantic", "evidence_chain",
    "relationship_strength", "blast_radius", "blast_radius_exposure",
    "severity", "rank", "pillar_id", "Result",
}

# Numeric score columns (for conditional formatting)
SCORE_COLS = {"claim_priority_score", "relationship_strength"}
BLAST_RADIUS_COLS = {"blast_radius", "blast_radius_exposure"}
HUB_TAG_COLS = {"hub_classification_tag", "hub_tag"}

# Columns containing claim_id references (for hyperlinking back to Claim Register)
# Single-value: cell contains exactly one claim_id
CLAIM_ID_SINGLE_COLS = {"claim_id", "claim_id_a", "claim_id_b", "hub_claim_id"}
# Multi-value: cell contains comma-separated claim_ids — link to the first one
CLAIM_ID_MULTI_COLS = {"linked_claims", "cascade_claims", "supporting_claim_ids", "hub_claims_linked"}
CLAIM_ID_ALL_COLS = CLAIM_ID_SINGLE_COLS | CLAIM_ID_MULTI_COLS

HYPERLINK_FONT = Font(name="Arial", size=10, color="1558B0", underline="single")
HYPERLINK_FONT_CENTER = Font(name="Arial", size=10, color="1558B0", underline="single")


def build_claim_row_map(wb):
    """Scan the Claim Register sheet and return {claim_id: row_number}."""
    claim_map = {}
    if "Claim Register" not in wb.sheetnames:
        return claim_map

    ws = wb["Claim Register"]
    headers = find_header_map(ws)
    cid_col = headers.get("claim_id")
    if not cid_col:
        return claim_map

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=cid_col).value
        if val:
            claim_map[str(val).strip()] = r

    return claim_map


def apply_claim_hyperlinks(ws, claim_map):
    """Add internal hyperlinks from claim_id cells to the Claim Register."""
    if not claim_map:
        return

    headers = find_header_map(ws)
    max_row = ws.max_row

    # Find the claim_id column in the Claim Register for the target cell reference
    # We link to column A (claim_id) of the target row
    for col_name, col_idx in headers.items():
        if col_name not in CLAIM_ID_ALL_COLS:
            continue

        is_multi = col_name in CLAIM_ID_MULTI_COLS

        for r in range(2, max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            raw = cell.value
            if not raw:
                continue

            raw_str = str(raw).strip()
            if not raw_str:
                continue

            if is_multi:
                # Extract first claim_id from comma-separated list
                parts = [p.strip() for p in raw_str.split(",") if p.strip()]
                target_id = None
                for part in parts:
                    if part in claim_map:
                        target_id = part
                        break
                if not target_id:
                    continue
            else:
                target_id = raw_str
                if target_id not in claim_map:
                    continue

            target_row = claim_map[target_id]
            # Internal hyperlink: #'Sheet Name'!Cell
            cell.hyperlink = f"#'Claim Register'!A{target_row}"
            cell.font = HYPERLINK_FONT


def style_data_sheet(ws, claim_map=None):
    """Apply professional formatting to a data sheet."""
    if ws.max_row < 1:
        return

    headers = find_header_map(ws)
    max_col = ws.max_column
    max_row = ws.max_row

    # ── Column widths ────────────────────────────────────────────
    for header_name, col_idx in headers.items():
        width = COLUMN_WIDTHS.get(header_name, 18)
        ws.column_dimensions[col_letter(col_idx)].width = width

    # ── Header row ───────────────────────────────────────────────
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = CELL_BORDER
    ws.row_dimensions[1].height = 28

    # ── Data rows ────────────────────────────────────────────────
    for r in range(2, max_row + 1):
        is_alt = (r % 2 == 0)
        fill = ALT_ROW_FILL if is_alt else WHITE_FILL

        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = CELL_BORDER
            cell.fill = fill

            # Determine column name
            header_cell = ws.cell(row=1, column=c).value
            col_name = str(header_cell).strip() if header_cell else ""

            if col_name in CENTER_COLS:
                cell.alignment = DATA_ALIGNMENT_CENTER
            else:
                cell.alignment = DATA_ALIGNMENT

            cell.font = DATA_FONT

    # ── Freeze panes ─────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ──────────────────────────────────────────────
    if max_row >= 2:
        ws.auto_filter.ref = f"A1:{col_letter(max_col)}{max_row}"

    # ── Conditional formatting ───────────────────────────────────
    _apply_conditional_formatting(ws, headers, max_row)

    # ── Claim ID hyperlinks ──────────────────────────────────────
    if claim_map and ws.title != "Claim Register":
        apply_claim_hyperlinks(ws, claim_map)


def _apply_conditional_formatting(ws, headers, max_row):
    """Apply conditional formatting rules based on column content."""
    if max_row < 2:
        return

    # ── Blast radius: red >= 5, amber >= 3 ───────────────────────
    for col_name in BLAST_RADIUS_COLS:
        if col_name not in headers:
            continue
        cl = col_letter(headers[col_name])
        rng = f"{cl}2:{cl}{max_row}"

        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["5"],
                font=Font(name="Arial", size=10, bold=True, color=RED_TEXT),
                fill=PatternFill("solid", fgColor=RED_BG),
            ),
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="between",
                formula=["3", "4"],
                font=Font(name="Arial", size=10, bold=True, color=AMBER_TEXT),
                fill=PatternFill("solid", fgColor=AMBER_BG),
            ),
        )

    # ── Claim priority score: red >= 0.80, amber >= 0.60 ────────
    for col_name in SCORE_COLS:
        if col_name not in headers:
            continue
        cl = col_letter(headers[col_name])
        rng = f"{cl}2:{cl}{max_row}"

        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["0.80"],
                font=Font(name="Arial", size=10, bold=True, color=RED_TEXT),
                fill=PatternFill("solid", fgColor=RED_BG),
            ),
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="between",
                formula=["0.60", "0.79"],
                font=Font(name="Arial", size=10, bold=True, color=AMBER_TEXT),
                fill=PatternFill("solid", fgColor=AMBER_BG),
            ),
        )

    # ── Hub classification tags: color by tag ────────────────────
    for col_name in HUB_TAG_COLS:
        if col_name not in headers:
            continue
        cl = col_letter(headers[col_name])
        rng = f"{cl}2:{cl}{max_row}"

        for tag, (bg, text) in HUB_COLORS.items():
            ws.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[f'EXACT({cl}2,"{tag}")'],
                    font=Font(name="Arial", size=10, bold=True, color=text),
                    fill=PatternFill("solid", fgColor=bg),
                ),
            )

    # ── Gate status: color pass/fail ─────────────────────────────
    if "gate_status" in headers:
        cl = col_letter(headers["gate_status"])
        rng = f"{cl}2:{cl}{max_row}"

        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'UPPER({cl}2)="PASS"'],
                font=Font(name="Arial", size=10, bold=True, color=GREEN_TEXT),
                fill=PatternFill("solid", fgColor=GREEN_BG),
            ),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'UPPER({cl}2)="FAIL"'],
                font=Font(name="Arial", size=10, bold=True, color=RED_TEXT),
                fill=PatternFill("solid", fgColor=RED_BG),
            ),
        )

    # ── Absence Claim highlighting ───────────────────────────────
    if "claim_type" in headers:
        cl = col_letter(headers["claim_type"])
        # Highlight entire row if claim_type is "Absence Claim"
        full_rng = f"A2:{col_letter(ws.max_column)}{max_row}"
        ws.conditional_formatting.add(
            full_rng,
            FormulaRule(
                formula=[f'${cl}2="Absence Claim"'],
                font=Font(name="Arial", size=10, italic=True, color=PURPLE_TEXT),
                fill=PatternFill("solid", fgColor=PURPLE_BG),
            ),
        )

    # ── Self-Audit / Export Validation: pass/fail result column ──
    if "Result" in headers:
        cl = col_letter(headers["Result"])
        rng = f"{cl}2:{cl}{max_row}"

        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'UPPER({cl}2)="PASS"'],
                font=Font(name="Arial", size=10, bold=True, color=GREEN_TEXT),
                fill=PatternFill("solid", fgColor=GREEN_BG),
            ),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'OR(UPPER({cl}2)="FAIL",UPPER({cl}2)="FALSE")'],
                font=Font(name="Arial", size=10, bold=True, color=RED_TEXT),
                fill=PatternFill("solid", fgColor=RED_BG),
            ),
        )


# ─── Main ─────────────────────────────────────────────────────────

EXPECTED_SHEETS = {"Claim Register", "Underwriting Gates", "Thesis Pillars"}


def main():
    if len(sys.argv) < 2 or sys.argv[1] in ("--help", "-h"):
        print(__doc__.strip())
        sys.exit(0)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else input_path

    if not os.path.exists(input_path):
        print(f"ERROR: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    try:
        wb = load_workbook(input_path)
    except Exception as e:
        print(f"ERROR: Cannot read workbook: {e}", file=sys.stderr)
        sys.exit(1)

    # Validate it looks like a Dataset D
    sheet_names = set(wb.sheetnames)
    if not EXPECTED_SHEETS.issubset(sheet_names):
        missing = EXPECTED_SHEETS - sheet_names
        print(f"ERROR: Not a Dataset D workbook. Missing sheets: {missing}", file=sys.stderr)
        sys.exit(2)

    try:
        # Build claim_id → row map from Claim Register (used for hyperlinking)
        claim_map = build_claim_row_map(wb)

        for name in wb.sheetnames:
            ws = wb[name]
            if name == "README":
                style_readme(ws)
            else:
                style_data_sheet(ws, claim_map)

        wb.save(output_path)
        print(f"OK: {output_path} ({os.path.getsize(output_path) / 1024:.1f} KB)")

    except Exception as e:
        print(f"ERROR: Formatting failed: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(3)


if __name__ == "__main__":
    main()
